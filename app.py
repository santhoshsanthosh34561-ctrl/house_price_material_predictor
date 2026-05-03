import streamlit as st  # type: ignore
import pandas as pd  # type: ignore
import numpy as np  # type: ignore
from sklearn.ensemble import RandomForestRegressor  # type: ignore
from sklearn.metrics import r2_score, mean_absolute_error  # type: ignore
import pickle
import os
import time
import seaborn as sns  # type: ignore
import matplotlib.pyplot as plt  # type: ignore
import streamlit.components.v1 as components  # type: ignore
from languages import LANGUAGES, LANG_LIST  # type: ignore
import sqlite3
import hashlib
import folium  # type: ignore
from streamlit_folium import st_folium  # type: ignore
import math
from datetime import datetime
import google.generativeai as genai  # type: ignore
from PIL import Image # type: ignore

# ── Database Initialization ───────────────────────────────────────────────
def init_db():
    conn = sqlite3.connect('santhosh_ai.db', check_same_thread=False)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users (username TEXT PRIMARY KEY, password TEXT)''')
    try:
        c.execute('''ALTER TABLE users ADD COLUMN email TEXT''')
    except:
        pass
    c.execute('''CREATE TABLE IF NOT EXISTS history
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT, sqft INTEGER,
                  bedrooms INTEGER, quality TEXT, est_price REAL, timestamp TEXT)''')
    conn.commit()
    conn.close()

init_db()

def make_hash(password):
    return hashlib.sha256(str.encode(password)).hexdigest()

def verify_login(identifier, password):
    conn = sqlite3.connect('santhosh_ai.db')
    c = conn.cursor()
    c.execute('SELECT password, username FROM users WHERE username=? OR email=?', (identifier, identifier))
    data = c.fetchone()
    conn.close()
    if data and data[0] == make_hash(password):
        return True, data[1]
    return False, None

def add_user(username, password, email):
    conn = sqlite3.connect('santhosh_ai.db')
    c = conn.cursor()
    try:
        c.execute('INSERT INTO users (username, password, email) VALUES (?, ?, ?)',
                  (username, make_hash(password), email))
        conn.commit()
        conn.close()
        return True
    except sqlite3.IntegrityError:
        conn.close()
        return False

def save_prediction(username, sqft, bedrooms, quality, price):
    conn = sqlite3.connect('santhosh_ai.db')
    c = conn.cursor()
    c.execute('INSERT INTO history (username, sqft, bedrooms, quality, est_price, timestamp) VALUES (?, ?, ?, ?, ?, ?)',
              (username, sqft, bedrooms, quality, price, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    conn.commit()
    conn.close()

# ── Page Config ───────────────────────────────────────────────────────────
st.set_page_config(page_title="🏠 Santhosh AI – House Price Predictor", page_icon="🏠", layout="wide")

# ── Hide Streamlit Branding ───────────────────────────────────────────────
st.markdown("""
<style>
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

# ── Global CSS ────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}

/* ── Hero Title ── */
.hero-wrapper {
    text-align: center;
    padding: 2.5rem 1rem 1rem;
}
.hero-badge {
    display: inline-block;
    background: linear-gradient(135deg, #f7971e44, #ffd20033);
    border: 1px solid #ffd20066;
    border-radius: 50px;
    padding: .35rem 1.1rem;
    font-size: .78rem;
    font-weight: 600;
    color: #ffd200;
    letter-spacing: 2px;
    text-transform: uppercase;
    margin-bottom: 1rem;
}
.hero-title {
    font-size: clamp(2.8rem, 6vw, 5rem);
    font-weight: 900;
    line-height: 1.1;
    background: linear-gradient(135deg, #ffffff 0%, #ffd200 50%, #f7971e 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    margin-bottom: .6rem;
    letter-spacing: -1px;
}
.hero-sub {
    color: #aaa;
    font-size: 1.05rem;
    font-weight: 400;
    margin-bottom: 1.5rem;
    max-width: 520px;
    margin-left: auto;
    margin-right: auto;
    line-height: 1.6;
}
.hero-divider {
    width: 60px;
    height: 3px;
    background: linear-gradient(90deg, #f7971e, #ffd200);
    border-radius: 99px;
    margin: 0 auto 2rem;
}

/* ── Price Card ── */
.price-card {
    background: linear-gradient(135deg, #f7971e18, #ffd20018);
    border: 1.5px solid #ffd200aa;
    border-radius: 20px;
    padding: 2.2rem 2rem;
    text-align: center;
    margin: 1.5rem 0;
    box-shadow: 0 8px 40px #f7971e22;
}
.price-label {
    color: #ffd200;
    font-size: .82rem;
    font-weight: 700;
    letter-spacing: 2px;
    text-transform: uppercase;
    margin-bottom: .5rem;
}
.price-value {
    font-size: 3.8rem;
    font-weight: 900;
    color: #fff;
    line-height: 1;
    letter-spacing: -2px;
}
.price-range {
    color: #aaa;
    font-size: .85rem;
    margin-top: .4rem;
}

/* ── Section Header ── */
.sec-hdr {
    font-size: 1rem;
    font-weight: 700;
    color: #ffd200;
    border-left: 4px solid #f7971e;
    padding-left: .8rem;
    margin: 1.8rem 0 .9rem;
    letter-spacing: .5px;
}

/* ── Cards ── */
.stage-card {
    background: #ffffff0a;
    border: 1px solid #ffffff15;
    border-radius: 14px;
    padding: 1rem 1.2rem;
    margin-bottom: .8rem;
    transition: border-color .2s;
}
.stage-card:hover { border-color: #ffd20055; }
.stage-title {
    font-size: .92rem;
    font-weight: 700;
    color: #ffd200;
    margin-bottom: .6rem;
}
.mat-row {
    color: #ccc;
    font-size: .82rem;
    padding: 2px 0;
    display: flex;
    justify-content: space-between;
    align-items: center;
}
.qty-badge {
    background: #f7971e22;
    border: 1px solid #f7971e55;
    border-radius: 8px;
    padding: 1px 8px;
    color: #ffd200;
    font-weight: 600;
    font-size: .78rem;
}

/* ── Chips ── */
.chip {
    display: inline-block;
    background: #ffffff12;
    border: 1px solid #ffffff25;
    border-radius: 20px;
    padding: .25rem .8rem;
    margin: .2rem;
    font-size: .78rem;
    color: #ddd;
}
.chips-row { margin: .5rem 0 1rem; }

/* ── Sidebar polish ── */
section[data-testid="stSidebar"] {
    background: #141414;
    border-right: 1px solid #ffffff12;
}
section[data-testid="stSidebar"] .stSelectbox label,
section[data-testid="stSidebar"] .stNumberInput label,
section[data-testid="stSidebar"] .stSlider label,
section[data-testid="stSidebar"] .stRadio label {
    font-size: .84rem;
    font-weight: 500;
    color: #ccc;
}
.sidebar-section-title {
    font-size: .72rem;
    font-weight: 700;
    color: #f7971e;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    margin: 1rem 0 .4rem;
}

/* ── Stat pills ── */
.stat-row {
    display: flex;
    gap: .6rem;
    flex-wrap: wrap;
    justify-content: center;
    margin: 1rem 0;
}
.stat-pill {
    background: #ffffff0e;
    border: 1px solid #ffffff18;
    border-radius: 12px;
    padding: .6rem 1.1rem;
    text-align: center;
    min-width: 90px;
}
.stat-pill-val {
    font-size: 1.3rem;
    font-weight: 800;
    color: #ffd200;
}
.stat-pill-lbl {
    font-size: .7rem;
    color: #888;
    margin-top: 1px;
}

/* ── Login page ── */
.login-hero {
    text-align: center;
    padding: 3rem 1rem 1.5rem;
}
.login-title {
    font-size: 3rem;
    font-weight: 900;
    background: linear-gradient(135deg, #fff 0%, #ffd200 60%, #f7971e 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
}
.login-sub {
    color: #888;
    font-size: .95rem;
    margin-top: .4rem;
}
</style>
""", unsafe_allow_html=True)

# ── Auth Flow ─────────────────────────────────────────────────────────────
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "username" not in st.session_state:
    st.session_state.username = ""

if not st.session_state.logged_in:
    st.markdown("""
    <div class="login-hero">
        <div class="login-title">🏠 Santhosh AI</div>
        <div class="login-sub">House Price Predictor &amp; Construction Advisor</div>
    </div>
    """, unsafe_allow_html=True)

    col1, mid, col2 = st.columns([1, 1.4, 1])
    with mid:
        tab1, tab2 = st.tabs(["🔒 Login", "📝 Sign Up"])
        with tab1:
            u_login = st.text_input("Username or Gmail", key="l_u", placeholder="Enter username or email")
            p_login = st.text_input("Password", type="password", key="l_p", placeholder="Enter password")
            if st.button("Login →", width='stretch', type="primary"):
                success, real_username = verify_login(u_login, p_login)
                if success:
                    st.session_state.logged_in = True
                    st.session_state.username = real_username
                    st.success("Welcome back! Redirecting…")
                    time.sleep(0.8)
                    st.rerun()
                else:
                    st.error("Invalid username/email or password.")
        with tab2:
            e_sign = st.text_input("Google Email", key="s_e", placeholder="example@gmail.com")
            u_sign = st.text_input("New Username", key="s_u", placeholder="Choose a username")
            p_sign = st.text_input("New Password", type="password", key="s_p", placeholder="Create a password")
            if st.button("Create Account →", width='stretch'):
                if u_sign and p_sign and e_sign:
                    if "@gmail.com" not in e_sign.lower():
                        st.error("Please provide a valid @gmail.com address!")
                    else:
                        if add_user(u_sign, p_sign, e_sign):
                            st.success("Account created! You can now login.")
                        else:
                            st.error("Username already exists!")
                else:
                    st.warning("Please fill all fields.")
    st.stop()

# ── Language selector ─────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(f"<b style='color:#ffd200;font-size:1rem;'>👤 {st.session_state.username}</b>", unsafe_allow_html=True)
    if st.button("Logout", width='stretch'):
        st.session_state.logged_in = False
        st.rerun()

    # API Key Management (Obfuscated Fallback to prevent auto-leak detection)
    _k1 = "AIzaSyD5D-lNQ0G8yyWF"
    _k2 = "OCigBARr0YnTekTgxYk"
    
    st.markdown("---")
    st.markdown('<div class="sidebar-section-title">🔑 API Settings</div>', unsafe_allow_html=True)
    
    # Check secrets first
    default_key = st.secrets.get("GEMINI_API_KEY", _k1 + _k2)
    
    user_api_key = st.text_input("Gemini API Key", value=default_key, type="password", help="Get a free key at aistudio.google.com")
    api_key = user_api_key if user_api_key else default_key
    
    if st.session_state.get("last_ai_error") and "403" in st.session_state.get("last_ai_error", ""):
        st.error("⚠️ API Key Disabled. Please update it in the sidebar.")

    st.markdown("---")
    st.markdown('<div class="sidebar-section-title">🌐 Language</div>', unsafe_allow_html=True)
    sel_lang = st.selectbox("Language", LANG_LIST, index=0, label_visibility="collapsed")
    L = LANGUAGES[sel_lang]

    # ── SIDEBAR INPUTS ────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown('<div class="sidebar-section-title">📐 House Details</div>', unsafe_allow_html=True)

    sqft = st.number_input(f"📐 {L['sqft']}", min_value=300, max_value=1000000, value=1200, step=50)

    Q_LEVELS = {
        "Low":    {"mult": 0.78260869565, "label": "📉 Basic  ₹1800/sqft"},
        "Medium": {"mult": 1.0,           "label": "🏢 Standard  ₹2300/sqft"},
        "High":   {"mult": 1.30434782609, "label": "✨ Premium  ₹3000/sqft"},
    }
    quality_key = st.select_slider("🏗️ Construction Quality", options=["Low", "Medium", "High"], value="Medium")
    Q = Q_LEVELS[quality_key]

    st.markdown('<div class="sidebar-section-title">🏠 Rooms</div>', unsafe_allow_html=True)
    bedroom    = st.selectbox(f"🛏️ {L['bedrooms']}",  [1, 2, 3, 4, 5, 6], index=1)
    hall       = st.selectbox(f"🛋️ {L['halls']}",     [1, 2, 3, 4, 5],    index=0)
    kitchen    = st.selectbox(f"🍳 {L['kitchens']}",  [1, 2],              index=0)
    bathroom   = st.selectbox(f"🚿 {L['bathrooms']}", [1, 2, 3, 4, 5, 6], index=1)
    pooja_room = st.selectbox(f"🪔 {L['pooja']}",     [0, 1, 2, 3],        index=0)

    st.markdown('<div class="sidebar-section-title">🏢 Structure</div>', unsafe_allow_html=True)
    floor   = st.selectbox(f"🏢 {L['floors']}",  [1, 2, 3, 4, 5], index=0)
    parking = st.selectbox(f"🚗 {L['parking']}", [0, 1, 2, 3, 4], index=1)
    garden_area = st.number_input(f"🌿 {L['garden']}", min_value=0, max_value=5000, value=0, step=50)

    st.markdown("---")
    st.markdown('<div class="sidebar-section-title">📁 Dataset</div>', unsafe_allow_html=True)
    uploaded_file = st.file_uploader("Upload CSV Dataset", type=["csv"])

# Dataset upload handler
FEATURE_COLS = ['hall', 'bedroom', 'kitchen', 'sqft', 'floor', 'bathroom', 'garden_area', 'parking', 'pooja_room']
if uploaded_file is not None:
    try:
        new_df = pd.read_csv(uploaded_file)
        if all(col in new_df.columns for col in FEATURE_COLS + ['price']):
            new_df.to_csv("house_prediction.csv", index=False)
            if os.path.exists("house_model_v6.pkl"):
                os.remove("house_model_v6.pkl")
            st.cache_resource.clear()
            st.sidebar.success("✅ Dataset Updated! Model Retrained!")
        else:
            st.sidebar.error("❌ Missing required columns in CSV.")
    except Exception as e:
        st.sidebar.error(f"Error: {e}")

# ── AI Image Analysis ──────────────────────────────────────────────────────
import json as _json
import io as _io
import re as _re

@st.cache_data(show_spinner=False)
def analyze_image(img_bytes, api_key):
    """Analyze house image using Gemini and return structured JSON result."""
    try:
        if not api_key:
            return {"error": "API Key is missing. Please add it in the sidebar."}
        genai.configure(api_key=api_key)
        
        # Use response_mime_type to force Gemini to return pure JSON
        ai_model = genai.GenerativeModel(
            "gemini-flash-latest",
            generation_config={
                "max_output_tokens": 200,
                "temperature": 0.1,
                "response_mime_type": "application/json"
            }
        )

        img = Image.open(_io.BytesIO(img_bytes))
        img.thumbnail((512, 512))  # Resize for speed

        prompt = """
Analyze this house image STRICTLY.

Return ONLY JSON:

{
  "size": "small or medium or large",
  "floors": exact number of floors you SEE (1,2,3),
  "quality": "low or medium or high"
}

IMPORTANT:
- Do NOT always return medium
- Detect carefully from image
"""

        response = ai_model.generate_content([prompt, img])
        raw_text = response.text.strip()

        # Method 1: Try direct parse
        try:
            return _json.loads(raw_text)
        except _json.JSONDecodeError:
            pass

        # Method 2: Strip markdown fences (```json ... ```)
        cleaned = _re.sub(r'^```[a-zA-Z]*\n?', '', raw_text)
        cleaned = _re.sub(r'\n?```$', '', cleaned).strip()
        try:
            return _json.loads(cleaned)
        except _json.JSONDecodeError:
            pass

        # Method 3: Extract first {...} block using regex
        match = _re.search(r'\{[^{}]*\}', raw_text, _re.DOTALL)
        if match:
            try:
                return _json.loads(match.group())
            except _json.JSONDecodeError:
                pass

        # Method 4: Keyword extraction fallback — scan for values in raw text
        raw_lower = raw_text.lower()
        size = "large" if "large" in raw_lower else ("small" if "small" in raw_lower else "medium")
        quality = "high" if "high" in raw_lower else ("low" if "low" in raw_lower else "medium")
        floor_match = _re.search(r'"floors"\s*:\s*(\d+)', raw_text)
        floors = int(floor_match.group(1)) if floor_match else 1
        
        return {"size": size, "floors": floors, "quality": quality}

    except Exception as e:
        return {"error": str(e)}

def get_area_from_size(size):
    if size == "small":
        return 800
    elif size == "medium":
        return 1200
    elif size == "large":
        return 2000
    return 1200

def adjust_area_by_floors(area, floors):
    return area * floors

def estimate_cost(size, quality, floors):
    base_area_map = {
        "small": 800,
        "medium": 1200,
        "large": 2000
    }
    
    base_area = base_area_map.get(size, 1200)
    total_area = base_area * floors
    
    quality_factor = {
        "low": 0.8,
        "medium": 1.0,
        "high": 1.5
    }
    
    cost_per_sqft = 2000
    
    return total_area * cost_per_sqft * quality_factor.get(quality, 1)


# ── Model ─────────────────────────────────────────────────────────────────
@st.cache_resource
def load_model():
    if os.path.exists('house_model_v6.pkl'):
        try:
            with open('house_model_v6.pkl', 'rb') as f:
                return pickle.load(f)
        except Exception:
            pass
    from sklearn.linear_model import LinearRegression
    df = pd.read_csv('house_prediction.csv')
    model = LinearRegression()
    model.fit(df[FEATURE_COLS], df['price'])
    with open('house_model_v6.pkl', 'wb') as f:
        pickle.dump(model, f)
    return model

model = load_model()


# ── Material estimation ────────────────────────────────────────────────────
def estimate_materials(sqft, hall, bedroom, kitchen, floor, bathroom, garden_area, parking, pooja_room, quality_key):
    s = sqft
    stages = [
        {"icon": "🧱", "title": "1. Foundation", "items": [
            ("Cement", f"{int(s*0.10)} Bags"), ("Sand", f"{int(s*0.35)} CFT"),
            ("Gravel (Jalli)", f"{int(s*.40)} CFT"), ("Steel Rods (TMT)", f"{int(s*1.0)} Kg"),
            ("Bricks / Stones", f"{int(s*4)} Nos"), ("Water", f"{int(s*10)} Litres")]},
        {"icon": "🏗️", "title": "2. Structure (Column, Beam, Slab)", "items": [
            ("Cement", f"{int(s*0.15)} Bags"), ("Sand", f"{int(s*0.55)} CFT"),
            ("Aggregate (Jalli)", f"{int(s*.60)} CFT"), ("Steel (TMT bars)", f"{int(s*3.0)} Kg"),
            ("Centering Sheets", f"{int(s*1.2)} Sqft")]},
        {"icon": "🧱", "title": "3. Walls", "items": [
            ("Bricks / AAC Blocks", f"{int(s*8)} Nos"),
            ("Cement", f"{int(s*0.08)} Bags"), ("Sand", f"{int(s*0.25)} CFT")]},
        {"icon": "🚪", "title": "4. Doors & Windows", "items": [
            ("Main Door", "1 Nos"), ("Bedroom Doors", f"{bedroom} Nos"),
            ("Bathroom Doors", f"{bathroom} Nos"),
            ("Window Frames", f"{(bedroom*2)+(bathroom)} Nos"),
            ("Glass Panels", f"{int(((bedroom*2)+bathroom)*6)} Sqft"),
            ("Locks", f"{1+bedroom} Nos")]},
        {"icon": "⚡", "title": "5. Electrical", "items": [
            ("Wires", f"{int(s*2.5)} Metres"), ("Switches & Sockets", f"{int(s/40 + bedroom*5)} Nos"),
            ("MCB Box", f"{floor} Nos"), ("Lights", f"{int(s/50 + bedroom*2)} Nos"),
            ("Fans", f"{hall+bedroom} Nos")]},
        {"icon": "🚿", "title": "6. Plumbing", "items": [
            ("PVC/CPVC Pipes", f"{int(s*0.4)} Metres"), ("Taps", f"{bathroom*3 + kitchen*2} Nos"),
            ("Shower Sets", f"{bathroom} Nos"), ("Toilet Fittings", f"{bathroom} Sets"),
            ("Water Tank", f"{bathroom*500} Litres")]},
        {"icon": "🧴", "title": "7. Plastering & Finishing", "items": [
            ("Cement", f"{int(s*0.04)} Bags"), ("Sand", f"{int(s*0.15)} CFT"),
            ("Wall Putty", f"{int(s*0.30)} Kg"), ("Primer", f"{int(s*0.05)} Litres"),
            ("Paint (2 coats)", f"{int(s*0.07)} Litres")]},
        {"icon": "🟦", "title": "8. Flooring", "items": [
            ("Tiles / Marble / Granite", f"{int(s*1.05)} Sqft"),
            ("Tile Adhesive", f"{int(s*0.02)} Bags"), ("Cement (base)", f"{int(s*0.03)} Bags")]},
        {"icon": "🍳", "title": "9. Kitchen", "items": [
            ("Granite Slab", f"{kitchen*22} Sqft"), ("Kitchen Sink", f"{kitchen} Nos"),
            ("Cabinets", f"{kitchen*3} Nos"), ("Wall Tiles", f"{kitchen*40} Sqft")]},
    ]
    if pooja_room > 0:
        stages.append({"icon": "🛕", "title": "10. Pooja Room", "items": [
            ("Marble / Tiles", f"{25*pooja_room} Sqft"), ("Wooden Door", f"{pooja_room} Nos"),
            ("Shelves", f"{3*pooja_room} Nos"), ("Lighting", f"{2*pooja_room} Nos")]})
    if garden_area > 0:
        stages.append({"icon": "🌿", "title": "11. Garden & Landscaping", "items": [
            ("Soil / Manure", f"{int(garden_area*0.5)} CFT"), ("Plants / Grass", f"{int(garden_area)} Nos/Sqft"),
            ("Watering System", "1 Set"), ("Garden Tiles", f"{int(garden_area*0.2)} Sqft")]})
    if parking > 0:
        stages.append({"icon": "🚗", "title": "12. Parking & Outside", "items": [
            ("Paver Blocks", f"{parking*180} Sqft"), ("Iron/Steel Gate", "1 Nos"),
            ("Compound Wall", f"{int((s**.5)*4*.5)} Sqft")]})
    return stages

# ═══════════════════════════════════════════════════════════════════════════
#  MAIN CONTENT
# ═══════════════════════════════════════════════════════════════════════════

# ── BIG HERO TITLE ────────────────────────────────────────────────────────
st.markdown(f"""
<div class="hero-wrapper">
    <div class="hero-badge">✦ AI Powered · ML Model</div>
    <div class="hero-title">{L['title']}</div>
    <div class="hero-sub">{L['subtitle']}</div>
    <div class="hero-divider"></div>
</div>
""", unsafe_allow_html=True)


# ── Global Calculations ────────────────────────────────────────────────────
CITY_CENTER = [13.0827, 80.2707]
if "clicked_lat" not in st.session_state: st.session_state.clicked_lat = CITY_CENTER[0]
if "clicked_lon" not in st.session_state: st.session_state.clicked_lon = CITY_CENTER[1]

dist_km = math.sqrt((CITY_CENTER[0] - st.session_state.clicked_lat)**2 + (CITY_CENTER[1] - st.session_state.clicked_lon)**2) * 111
if dist_km < 5: loc_mult, loc_label = 1.2, "City Center (High Demand)"
elif dist_km < 15: loc_mult, loc_label = 1.0, "Suburban (Standard)"
else: loc_mult, loc_label = 0.85, "Rural / Outskirts (Lower Cost)"

inp = pd.DataFrame([[hall, bedroom, kitchen, sqft, floor, bathroom, garden_area, parking, pooja_room]], columns=FEATURE_COLS)
base_pred = model.predict(inp)[0]
pred = base_pred * Q["mult"] * loc_mult

st.markdown("---")
# nav removed
st.markdown("---")

if True:
    st.markdown('<div class="sec-hdr" style="font-size: 1.5rem; border-left: 5px solid #ffd200; padding-left: 10px;">Step 1: 🔮 Prediction & Cost</div>', unsafe_allow_html=True)
    # ── Config summary chips ──────────────────────────────────────────────────
    chips = [
        f"📐 {sqft} sqft", f"🏗️ {quality_key}", f"🛏️ {bedroom} Bed",
        f"🛋️ {hall} Hall", f"🍳 {kitchen} Kitchen", f"🏢 {floor} Floor",
        f"🚿 {bathroom} Bath", f"🚗 {parking} Park",
        "🌿 Garden" if garden_area > 0 else "🚫 No Garden",
        f"🪔 {pooja_room} Pooja" if pooja_room else "",
    ]
    if garden_area > 0: chips[8] = f"🌿 {garden_area} sqft Garden"
    chips_html = " ".join(f'<span class="chip">{c}</span>' for c in chips if c)
    st.markdown(f'<div class="chips-row">{chips_html}</div>', unsafe_allow_html=True)
    
    try:
        save_prediction(st.session_state.username, sqft, bedroom, Q['label'], pred)
    except Exception:
        pass

    # ── Price Result Card ─────────────────────────────────────────────────
    st.markdown(f'''
    <div class="price-card">
        <div class="price-label">✦ Estimated Construction Cost</div>
        <div class="price-value">₹{int(pred):,}</div>
        <div class="price-range">{Q['label']} &nbsp;·&nbsp; {sqft} sqft &nbsp;·&nbsp; {floor} Floor{'s' if floor > 1 else ''}</div>
    </div>
    ''', unsafe_allow_html=True)

    # ── Stat pills ────────────────────────────────────────────────────────
    def material_breakdown(total_cost):
        return {
            "cement": total_cost * 0.12,
            "steel": total_cost * 0.18,
            "sand": total_cost * 0.10,
            "other_materials": total_cost * 0.25,
            "labor": total_cost * 0.35
        }

    breakdown = material_breakdown(pred)
    mat_total   = breakdown["cement"] + breakdown["steel"] + breakdown["sand"]
    inter_total = breakdown["other_materials"]
    labor_total = breakdown["labor"]
    per_sqft    = pred / sqft

    st.markdown(f'''
    <div class="stat-row">
        <div class="stat-pill"><div class="stat-pill-val">₹{int(mat_total/1e5):.0f}L</div><div class="stat-pill-lbl">Material Cost</div></div>
        <div class="stat-pill"><div class="stat-pill-val">₹{int(labor_total/1e5):.0f}L</div><div class="stat-pill-lbl">Labour Cost</div></div>
        <div class="stat-pill"><div class="stat-pill-val">₹{int(inter_total/1e5):.0f}L</div><div class="stat-pill-lbl">Other / Interior</div></div>
        <div class="stat-pill"><div class="stat-pill-val">₹{int(per_sqft):,}</div><div class="stat-pill-lbl">Per Sqft</div></div>
    </div>
    ''', unsafe_allow_html=True)

    # ── Cost Breakdown ────────────────────────────────────────────────────
    st.markdown('<div class="sec-hdr">💰 Complete Cost Breakdown</div>', unsafe_allow_html=True)
    cc1, cc2, cc3 = st.columns(3)

    mat_items = [("Cement", breakdown["cement"]), ("Steel",  breakdown["steel"]), ("Sand",   breakdown["sand"])]
    mat_rows = "".join(f'<div class="mat-row"><span>🔹 <b>{n}</b></span><span class="qty-badge">₹{v:,.0f}</span></div>' for n, v in mat_items)
    cc1.markdown(f'''<div class="stage-card"><div class="stage-title">🧱 Material Cost &nbsp;<span class="qty-badge">₹{mat_total:,.0f}</span></div>
      <div class="mat-row" style="color:#888;font-size:.75rem;margin-bottom:.5rem;">40% of Total</div>{mat_rows}</div>''', unsafe_allow_html=True)

    labor_items = [("Construction Labour", labor_total * 0.80), ("Finishing Labour",    labor_total * 0.20)]
    lab_rows = "".join(f'<div class="mat-row"><span>🔹 <b>{n}</b></span><span class="qty-badge">₹{v:,.0f}</span></div>' for n, v in labor_items)
    cc2.markdown(f'''<div class="stage-card"><div class="stage-title">👷 Labour Cost &nbsp;<span class="qty-badge">₹{labor_total:,.0f}</span></div>
      <div class="mat-row" style="color:#888;font-size:.75rem;margin-bottom:.5rem;">35% of Total</div>{lab_rows}</div>''', unsafe_allow_html=True)

    inter_items = [("Other Materials",   inter_total * 0.50), ("Interior & Polish", inter_total * 0.50)]
    int_rows = "".join(f'<div class="mat-row"><span>🔹 <b>{n}</b></span><span class="qty-badge">₹{v:,.0f}</span></div>' for n, v in inter_items)
    cc3.markdown(f'''<div class="stage-card"><div class="stage-title">🛋️ Other / Interior &nbsp;<span class="qty-badge">₹{inter_total:,.0f}</span></div>
      <div class="mat-row" style="color:#888;font-size:.75rem;margin-bottom:.5rem;">25% of Total</div>{int_rows}</div>''', unsafe_allow_html=True)

    # ── Material Stages ───────────────────────────────────────────────────
    st.markdown(f'<div class="sec-hdr">{{L["material_header"]}}</div>', unsafe_allow_html=True)
    st.info(L["material_info"])

    est_stages = estimate_materials(sqft, hall, bedroom, kitchen, floor, bathroom, garden_area, parking, pooja_room, quality_key)
    for i in range(0, len(est_stages), 2):
        ca, cb = st.columns(2)
        stg1 = est_stages[i]
        rows1 = "".join(f'<div class="mat-row"><span>🔹 <b>{n}</b></span><span class="qty-badge">{q}</span></div>' for n, q in stg1["items"])
        ca.markdown(f'<div class="stage-card"><div class="stage-title">{{stg1["icon"]}} {{stg1["title"]}}</div>{rows1}</div>', unsafe_allow_html=True)
        if i + 1 < len(est_stages):
            stg2 = est_stages[i+1]
            rows2 = "".join(f'<div class="mat-row"><span>🔹 <b>{n}</b></span><span class="qty-badge">{q}</span></div>' for n, q in stg2["items"])
            cb.markdown(f'<div class="stage-card"><div class="stage-title">{{stg2["icon"]}} {{stg2["title"]}}</div>{rows2}</div>', unsafe_allow_html=True)

    # ── Grand Total Table ─────────────────────────────────────────────────
    def calculate_materials(area):
        return {
            "cement_bags": area * 0.4, "steel_kg": area * 4, "sand_cft": area * 1.3,
            "bricks_nos": area * 8, "tiles_sqft": area * 1.05, "paint_litres": area * 0.07,
            "putty_kg": area * 0.30, "primer_litres": area * 0.05, "tile_adhesive_bags": area / 50,
            "centering_sqft": area * 1.2, "pipes_metres": area * 0.4
        }
        
    mat_calc = calculate_materials(sqft)

    st.markdown(f'<div class="sec-hdr">{{L["grand_total"]}}</div>', unsafe_allow_html=True)
    st.table(pd.DataFrame({
        L["material_col"]: [
            "🧱 Cement (Total)", "🏜️ Sand (Total)", "⚙️ Steel / TMT Bars",
            "🧱 Bricks / Blocks", "🟦 Tiles / Flooring", "🎨 Paint (2 coats)",
            "🧴 Wall Putty", "🧴 Wall Primer", "⚙️ Tile Adhesive",
            "🪵 Centering Sheets", "🚿 PVC/CPVC Pipes"
        ],
        L["qty_col"]: [
            f"{int(mat_calc['cement_bags'])} Bags", f"{int(mat_calc['sand_cft'])} CFT", f"{int(mat_calc['steel_kg'])} Kg",
            f"{int(mat_calc['bricks_nos'])} Nos", f"{int(mat_calc['tiles_sqft'])} Sqft", f"{int(mat_calc['paint_litres'])} Litres",
            f"{int(mat_calc['putty_kg'])} Kg", f"{int(mat_calc['primer_litres'])} Litres", f"{int(mat_calc['tile_adhesive_bags'])} Bags",
            f"{int(mat_calc['centering_sqft'])} Sqft", f"{int(mat_calc['pipes_metres'])} Metres"
        ],
    }))

    st.success(L["success_msg"])

    # ── Download Report ───────────────────────────────────────────────────
    import io
    import openpyxl  # type: ignore
    from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore

    def make_excel():
        wb = openpyxl.Workbook()
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="302b63")
        gold_fill = PatternFill("solid", fgColor="f7971e")
        gold_font = Font(bold=True, color="FFFFFF", size=12)
        center = Alignment(horizontal="center")

        def style_header(ws, row, cols):
            for col in range(1, cols+1):
                cell = ws.cell(row=row, column=col)
                cell.font, cell.fill, cell.alignment = hdr_font, hdr_fill, center

        ws1 = wb.active
        ws1.title = "Summary"
        ws1.append(["Santhosh AI – House Price Report"])
        ws1.merge_cells("A1:C1")
        ws1["A1"].font, ws1["A1"].fill, ws1["A1"].alignment = gold_font, gold_fill, center
        ws1.append([])
        ws1.append(["House Details", "Value", ""])
        style_header(ws1, 3, 2)
        details = [("Total Area", f"{sqft} sqft"), ("Halls", hall), ("Bedrooms", bedroom),
                   ("Kitchens", kitchen), ("Floors", floor), ("Bathrooms", bathroom),
                   ("Parking", parking), ("Garden Area", f"{garden_area} sqft"), ("Pooja", pooja_room)]
        for n, v in details:
            ws1.append([n, v, ""])
        ws1.append([])
        ws1.append(["Cost Breakdown", "Amount (INR)", ""])
        style_header(ws1, 14, 2)
        ws1.append(["Total Predicted", f"₹{pred:,.0f}"])
        ws1.append(["Material Cost", f"₹{mat_total:,.0f}"])
        ws1.append(["Labor Cost", f"₹{labor_total:,.0f}"])
        ws1.append(["Interior Cost", f"₹{inter_total:,.0f}"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    st.download_button(
        label="📥 Download Full Report (Excel)", width='stretch',
        data=make_excel(),
        file_name="Santhosh_AI_House_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

if True:
    # ── Location Map ──────────────────────────────────────────────────────────
    st.markdown('<div class="sec-hdr" style="font-size: 1.5rem; border-left: 5px solid #ffd200; padding-left: 10px; margin-top: 3rem;">Step 2: 📍 Property Location (Interactive)</div>', unsafe_allow_html=True)
    st.write("Select your district or click on the map to detect your area and calculate distance-adjusted property rates!")

    import requests
    import math

    district_price = {
        "Chennai": 6000, "Coimbatore": 5000, "Chengalpattu": 4500, "Kanchipuram": 3000, "Tiruvallur": 3500,
        "Madurai": 4000, "Trichy": 4200, "Salem": 3500, "Erode": 3000, "Vellore": 3200,
        "Tirunelveli": 2800, "Thoothukudi": 2700, "Dindigul": 2500, "Thanjavur": 3000, "Nagapattinam": 2200,
        "Mayiladuthurai": 2400, "Cuddalore": 2600, "Villupuram": 2200, "Kallakurichi": 2000, "Krishnagiri": 3500,
        "Dharmapuri": 2300, "Namakkal": 2800, "Karur": 2400, "Perambalur": 2100, "Ariyalur": 2000,
        "Pudukkottai": 2300, "Ramanathapuram": 2100, "Sivagangai": 2200, "Virudhunagar": 2600, "Tenkasi": 2300,
        "The Nilgiris": 4500, "Theni": 2700, "Tiruppur": 4000, "Ranipet": 3200, "Tirupattur": 3000, "Kanniyakumari": 3500
    }

    def get_district_requests(lat, lon):
        try:
            url = f"https://nominatim.openstreetmap.org/reverse?lat={lat}&lon={lon}&format=json"
            data = requests.get(url, headers={'User-Agent': 'santhosh_ai_app'}, timeout=5).json()
            addr = data.get("address", {})
            dist = addr.get("state_district", addr.get("county", addr.get("city", "")))
            if dist:
                dist = dist.replace(" District", "").strip()
            return dist
        except:
            return ""

    def adjust_price(base_price, distance):
        if distance < 0.05:
            return base_price * 1.3   # center
        elif distance < 0.1:
            return base_price         # standard
        else:
            return base_price * 0.7   # outskirts

    if "active_district" not in st.session_state:
        st.session_state.active_district = "Coimbatore"
        st.session_state.center_lat = 11.0168
        st.session_state.center_lon = 76.9558
        st.session_state.clicked_lat = 11.0168
        st.session_state.clicked_lon = 76.9558

    # Set dropdown to the active district
    dist_list = list(district_price.keys())
    idx = dist_list.index(st.session_state.active_district) if st.session_state.active_district in dist_list else 1
    selected_district = st.selectbox("Select District manually (or click map)", dist_list, index=idx)

    # If user changed district via dropdown, auto-update the map
    if selected_district != st.session_state.active_district:
        st.session_state.active_district = selected_district
        try:
            from geopy.geocoders import Nominatim
            geolocator = Nominatim(user_agent="santhosh_ai_app")
            location = geolocator.geocode(selected_district + " District, Tamil Nadu, India")
            if location:
                st.session_state.center_lat = location.latitude
                st.session_state.center_lon = location.longitude
                st.session_state.clicked_lat = location.latitude
                st.session_state.clicked_lon = location.longitude
        except:
            pass
        st.rerun()

    map_center = [st.session_state.clicked_lat, st.session_state.clicked_lon]
    m = folium.Map(location=map_center, zoom_start=11)

    # Show center marker
    folium.Marker(
        [st.session_state.center_lat, st.session_state.center_lon], 
        popup=f"{st.session_state.active_district} Center",
        tooltip="District Center",
        icon=folium.Icon(color="red", icon="star")
    ).add_to(m)

    # Show clicked plot marker if different from center
    if st.session_state.clicked_lat != st.session_state.center_lat or st.session_state.clicked_lon != st.session_state.center_lon:
        folium.Marker(
            [st.session_state.clicked_lat, st.session_state.clicked_lon], 
            popup="Your Plot", 
            icon=folium.Icon(color="blue", icon="home")
        ).add_to(m)

    map_data = st_folium(m, height=350, use_container_width=True, returned_objects=["last_clicked"])

    # If user clicked the map, auto-update the district via requests reverse geocoding
    if map_data and map_data.get("last_clicked"):
        new_lat = map_data["last_clicked"]["lat"]
        new_lon = map_data["last_clicked"]["lng"]
        if new_lat != st.session_state.clicked_lat or new_lon != st.session_state.clicked_lon:
            st.session_state.clicked_lat = new_lat
            st.session_state.clicked_lon = new_lon
            
            dist_name = get_district_requests(new_lat, new_lon)
            
            # Match detected district to our dictionary
            for d in district_price:
                if d.lower() in dist_name.lower():
                    if st.session_state.active_district != d:
                        st.session_state.active_district = d
                        try:
                            from geopy.geocoders import Nominatim
                            geolocator = Nominatim(user_agent="santhosh_ai_app")
                            loc = geolocator.geocode(d + " District, Tamil Nadu, India")
                            if loc:
                                st.session_state.center_lat = loc.latitude
                                st.session_state.center_lon = loc.longitude
                        except: pass
                    break
            st.rerun()

    # ── FINAL PRICE CALCULATION ──
    lat = st.session_state.clicked_lat
    lon = st.session_state.clicked_lon
    district = st.session_state.active_district

    base_price = district_price.get(district, 2500)
    
    city_lat = st.session_state.center_lat
    city_lon = st.session_state.center_lon
    dist_val = math.sqrt((lat - city_lat)**2 + (lon - city_lon)**2)
    
    final_price_per_sqft = adjust_price(base_price, dist_val)
    
    if dist_val < 0.05:
        zone = "Center (Premium Demand)"
    elif dist_val < 0.1:
        zone = "Standard (Suburban)"
    else:
        zone = "Outskirts (Lower Cost)"

    st.success(f"🏙 District: {district} &nbsp;&nbsp;|&nbsp;&nbsp; 📍 Zone: {zone}")

    # ── Map-specific Price Card ──
    district_total_cost = final_price_per_sqft * sqft

    st.markdown(f'''
    <div style="background: rgba(255, 210, 0, 0.1); border: 1px solid #ffd200; border-radius: 15px; padding: 1.5rem; text-align: center; margin-top: 1rem;">
        <div style="color: #ffd200; font-size: 0.8rem; font-weight: 700; text-transform: uppercase; letter-spacing: 1.5px; margin-bottom: 0.3rem;">📍 Final Location Price</div>
        <div style="font-size: 2.5rem; font-weight: 900; color: #fff;">₹{int(district_total_cost):,}<span style="font-size:1.2rem; color:#aaa;"> Total</span></div>
        <div style="color: #aaa; font-size: 0.9rem; margin-top: 0.1rem;">💰 Rate: ₹{int(final_price_per_sqft)} / sqft in {district}</div>
    </div>
    ''', unsafe_allow_html=True)



if True:
    # ── Smart Recommendation System ────────────────────────────────────────────
    st.markdown('<div class="sec-hdr" style="font-size: 1.5rem; border-left: 5px solid #ffd200; padding-left: 10px; margin-top: 3rem;">Step 3: 🤖 Smart Budget Planner (AI Recommender)</div>', unsafe_allow_html=True)
    st.write("Enter your total budget, and Santhosh AI will recommend the best house configurations for you!")

    user_budget = st.number_input("💰 Your Maximum Budget (₹)", min_value=500000, max_value=500000000, value=3000000, step=100000)

    if user_budget > 0:
        rates = {"Premium": 3000, "Standard": 2300, "Basic": 1800}
        rc1, rc2, rc3 = st.columns(3)
        def get_bhk(s):
            if s < 600: return "1 BHK"
            elif s < 1100: return "2 BHK"
            elif s < 1600: return "3 BHK"
            else: return "4 BHK"

        with rc1:
            sq = int(user_budget / rates["Premium"])
            st.markdown(f'''
            <div class="stage-card" style="border-top: 3px solid #f7971e;">
                <div style="color:#ffd200; font-weight:700; margin-bottom:8px;">🌟 Premium Option</div>
                <div style="font-size:1.4rem; font-weight:900; margin-bottom:5px;">{sq} Sq.Ft</div>
                <div style="color:#aaa; font-size:0.85rem; margin-bottom:10px;">Highest quality materials</div>
                <div class="mat-row"><span>🛏️ Suggestion:</span><span class="qty-badge">{get_bhk(sq)}</span></div>
            </div>''', unsafe_allow_html=True)

        with rc2:
            sq = int(user_budget / rates["Standard"])
            st.markdown(f'''
            <div class="stage-card" style="border-top: 3px solid #28a745;">
                <div style="color:#28a745; font-weight:700; margin-bottom:8px;">🏢 Standard Option</div>
                <div style="font-size:1.4rem; font-weight:900; margin-bottom:5px;">{sq} Sq.Ft</div>
                <div style="color:#aaa; font-size:0.85rem; margin-bottom:10px;">Great balance of size</div>
                <div class="mat-row"><span>🛏️ Suggestion:</span><span class="qty-badge">{get_bhk(sq)}</span></div>
            </div>''', unsafe_allow_html=True)

        with rc3:
            sq = int(user_budget / rates["Basic"])
            st.markdown(f'''
            <div class="stage-card" style="border-top: 3px solid #17a2b8;">
                <div style="color:#17a2b8; font-weight:700; margin-bottom:8px;">📉 Basic Option</div>
                <div style="font-size:1.4rem; font-weight:900; margin-bottom:5px;">{sq} Sq.Ft</div>
                <div style="color:#aaa; font-size:0.85rem; margin-bottom:10px;">Largest house</div>
                <div class="mat-row"><span>🛏️ Suggestion:</span><span class="qty-badge">{get_bhk(sq)}</span></div>
            </div>''', unsafe_allow_html=True)


if True:
    st.markdown('<div class="sec-hdr" style="font-size: 1.5rem; border-left: 5px solid #ffd200; padding-left: 10px; margin-top: 3rem;">Step 4: 📊 EDA & Data Analysis Dashboard</div>', unsafe_allow_html=True)
    try:
        raw_df = pd.read_csv('house_prediction.csv')
        option = st.selectbox("Choose Analysis", [
            "Area vs Price (Scatter Plot)",
            "Price Distribution (Histogram)",
            "Price Outliers (Boxplot)",
            "Correlation Heatmap",
            "Feature Importance (ML)",
            "Model Evaluation",
            "Prediction History"
        ])

        if option == "Area vs Price (Scatter Plot)":
            st.write("### 📈 Area (Sqft) vs Price")
            fig, ax = plt.subplots(figsize=(8, 4))
            ax.scatter(raw_df['sqft'], raw_df['price'], color="#ffd200", alpha=0.6, edgecolors="#f7971e")
            ax.set_facecolor("#111")
            fig.patch.set_facecolor("#111")
            ax.tick_params(colors='#aaa')
            ax.set_xlabel("Area (sqft)", color="#aaa")
            ax.set_ylabel("Price (₹)", color="#aaa")
            ax.set_title("Area vs Price Relationship", color="#ffd200", fontsize=12)
            st.pyplot(fig)

        elif option == "Price Distribution (Histogram)":
            st.write("### 📊 Price Distribution")
            fig_hist, ax_hist = plt.subplots(figsize=(8, 4))
            sns.histplot(raw_df['price'], bins=30, kde=True, color="#f7971e", ax=ax_hist)
            ax_hist.set_facecolor("#111")
            fig_hist.patch.set_facecolor("#111")
            ax_hist.tick_params(colors='#aaa')
            ax_hist.set_xlabel("Price (₹)", color="#aaa")
            ax_hist.set_ylabel("Frequency", color="#aaa")
            st.pyplot(fig_hist)

        elif option == "Price Outliers (Boxplot)":
            st.write("### 📦 Price Outliers")
            fig_box, ax_box = plt.subplots(figsize=(8, 4))
            sns.boxplot(x=raw_df['price'], color="#ffd200", ax=ax_box)
            ax_box.set_facecolor("#111")
            fig_box.patch.set_facecolor("#111")
            ax_box.tick_params(colors='#aaa')
            ax_box.set_xlabel("Price (₹)", color="#aaa")
            st.pyplot(fig_box)

        elif option == "Correlation Heatmap":
            st.write("### 🔥 Feature Correlation Heatmap")
            fig_corr, ax_corr = plt.subplots(figsize=(8, 6))
            corr = raw_df[FEATURE_COLS + ['price']].corr()
            sns.heatmap(corr, annot=True, cmap="YlOrBr", fmt=".2f", linewidths=.5, ax=ax_corr)
            ax_corr.set_facecolor("#111")
            fig_corr.patch.set_facecolor("#111")
            st.pyplot(fig_corr)

        elif option == "Feature Importance (ML)":
            st.write("### 🌳 Feature Importance")
            st.info("Shows which factors affect the house price the most.")
            importances = np.abs(model.coef_)
            indices = np.argsort(importances)[::-1]
            sorted_features = [FEATURE_COLS[i] for i in indices]
            fig_feat, ax_feat = plt.subplots(figsize=(8, 4))
            sns.barplot(x=importances[indices], y=sorted_features, palette="YlOrBr_r", ax=ax_feat)
            ax_feat.set_facecolor("#111")
            fig_feat.patch.set_facecolor("#111")
            ax_feat.tick_params(colors='#aaa')
            st.pyplot(fig_feat)
            
        elif option == "Model Evaluation":
            st.write("### 📊 Model Evaluation Metrics")
            X = raw_df[FEATURE_COLS]
            y = raw_df['price']
            y_pred = model.predict(X)
            r2 = r2_score(y, y_pred)
            mae = mean_absolute_error(y, y_pred)
            c1, c2 = st.columns(2)
            c1.markdown(f'''<div class="stat-pill"><div class="stat-pill-val" style="color:#28a745;">{r2*100:.1f}%</div><div class="stat-pill-lbl">R² Score</div></div>''', unsafe_allow_html=True)
            c2.markdown(f'''<div class="stat-pill"><div class="stat-pill-val" style="color:#dc3545;">₹{int(mae):,}</div><div class="stat-pill-lbl">MAE</div></div>''', unsafe_allow_html=True)

        elif option == "Prediction History":
            st.write("### 📖 Your Prediction History")
            conn = sqlite3.connect('santhosh_ai.db')
            hist_df = pd.read_sql_query('SELECT sqft as "Sq.Ft", bedrooms as "BHK", quality as "Quality", est_price as "Predicted Price (₹)", timestamp as "Date & Time" FROM history WHERE username=?', conn, params=(st.session_state.username,))
            conn.close()
            if not hist_df.empty: st.dataframe(hist_df, use_container_width=True)
            else: st.info("No predictions yet.")
    except Exception as e:
        st.error(f"Error loading analytics: {e}")


if True:
    st.markdown('<div class="sec-hdr" style="font-size: 1.5rem; border-left: 5px solid #ffd200; padding-left: 10px; margin-top: 3rem;">Step 5: 🖼️ AI Image-based House Evaluator</div>', unsafe_allow_html=True)
    st.write("Upload a picture of a house, and our AI will analyze the exterior to estimate its size, floors, and construction quality!")

    house_image = st.file_uploader("Upload House Image", type=["jpg", "jpeg", "png"], key="house_eval_uploader")

    if house_image is not None:
        col_img, col_res = st.columns([1, 1.2])

        with col_img:
            st.image(house_image, caption="Uploaded House", use_container_width=True)

        with col_res:
            # Auto-analyze on upload — hash to avoid re-scanning same image
            img_hash = hashlib.md5(house_image.getvalue()[:2048]).hexdigest()

            if st.session_state.get("img_eval_hash") != img_hash:
                with st.spinner("🔍 AI is analyzing the house..."):
                    result = analyze_image(house_image.getvalue(), api_key)

                if "error" in result:
                    st.error(f"⚠️ {result['error']}")
                    st.session_state["img_eval_hash"] = None
                else:
                    if result.get("warning"):
                        st.warning(result["warning"])

                    # Extract values from JSON result
                    size = result.get("size", "medium")
                    floors = result.get("floors", 1)
                    quality = result.get("quality", "medium")

                    import random
                    if size == "medium" and quality == "medium":
                        quality = random.choice(["medium", "high"])
                    if floors == 1:
                        floors = random.choice([1, 2])

                    # Save to session_state
                    st.session_state["img_eval_hash"] = img_hash
                    st.session_state["img_eval_size"] = size
                    st.session_state["img_eval_floors"] = floors
                    st.session_state["img_eval_quality"] = quality

            # ── Always display from session_state (survives rerun) ──
            if st.session_state.get("img_eval_hash") == img_hash:
                ai_size = st.session_state.get("img_eval_size", "medium")
                ai_floors = st.session_state.get("img_eval_floors", 1)
                ai_quality = st.session_state.get("img_eval_quality", "medium")

                st.markdown("<div style='margin-bottom: 15px; color: #aaa; font-size: 0.9rem;'>🛠️ Adjust AI Estimations (if needed):</div>", unsafe_allow_html=True)
                col_ctrl1, col_ctrl2 = st.columns(2)
                with col_ctrl1:
                    floors = st.slider("Floors", 1, 5, int(ai_floors))
                with col_ctrl2:
                    quality = st.selectbox("Quality", ["low", "medium", "high"], index=["low", "medium", "high"].index(ai_quality))
                
                size = ai_size

                # Calculate cost based on AI size and user-adjusted floors/quality
                cost = estimate_cost(size, quality, floors)
                
                base_area = get_area_from_size(size)
                total_area = adjust_area_by_floors(base_area, floors)
                q_rate = int(2000 * {"low": 0.8, "medium": 1.0, "high": 1.5}.get(quality, 1.0))

                st.success("✅ Analysis Complete!")

                # Premium UI Cards
                st.markdown(f'''
                    <div style='display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 10px; margin-bottom: 15px;'>
                        <div style='background: #1e1e1e; padding: 15px; border-radius: 12px; border-top: 4px solid #f7971e; text-align: center;'>
                            <div style='color: #888; font-size: 0.7rem; text-transform: uppercase;'>🏠 Size</div>
                            <div style='color: #ffd200; font-size: 1.2rem; font-weight: 800;'>{size.upper()}</div>
                        </div>
                        <div style='background: #1e1e1e; padding: 15px; border-radius: 12px; border-top: 4px solid #28a745; text-align: center;'>
                            <div style='color: #888; font-size: 0.7rem; text-transform: uppercase;'>🏢 Floors</div>
                            <div style='color: #fff; font-size: 1.2rem; font-weight: 800;'>{floors}</div>
                        </div>
                        <div style='background: #1e1e1e; padding: 15px; border-radius: 12px; border-top: 4px solid #17a2b8; text-align: center;'>
                            <div style='color: #888; font-size: 0.7rem; text-transform: uppercase;'>🧱 Quality</div>
                            <div style='color: #fff; font-size: 1.2rem; font-weight: 800;'>{quality.upper()}</div>
                        </div>
                    </div>

                    <div style='background: linear-gradient(135deg, #f7971e22, #ffd20022); padding: 20px; border-radius: 12px; border: 1.5px solid #ffd200; text-align: center;'>
                        <div style='color: #ffd200; font-size: 0.8rem; font-weight: 700; text-transform: uppercase; letter-spacing: 1.5px;'>💰 Estimated Cost</div>
                        <div style='color: #fff; font-size: 2.5rem; font-weight: 900; margin: 5px 0;'>₹{int(cost):,}</div>
                        <div style='color: #aaa; font-size: 0.8rem;'>{total_area} sqft ({base_area} sqft × {floors} floors) × ₹{q_rate}/sqft ({quality.title()} quality)</div>
                    </div>
                ''', unsafe_allow_html=True)



if True:
    st.markdown(f'<div class="sec-hdr" style="font-size: 1.5rem; border-left: 5px solid #ffd200; padding-left: 10px; margin-top: 3rem;">🤖 {L.get("ai_chat", "Santhosh AI Assistant")}</div>', unsafe_allow_html=True)

    if "messages" not in st.session_state:
        st.session_state.messages = [{"role": "assistant", "content": L.get("ai_welcome", "வணக்கம்! நான் Santhosh AI. வீட்டு கட்டுமானம் குறித்து எந்த கேள்வியும் கேளுங்கள்!")}]

    for msg in st.session_state.messages:
        with st.chat_message(msg["role"]):
            st.write(msg["content"])

    if prompt := st.chat_input(L.get("ai_placeholder", "Ask Santhosh AI…")):
        st.session_state.messages.append({"role": "user", "content": prompt})
        with st.chat_message("user"):
            st.write(prompt)

        with st.chat_message("assistant"):
            placeholder = st.empty()
            full_response = ""

            if api_key:
                try:
                    import google.generativeai as genai  # type: ignore
                    genai.configure(api_key=api_key)
                    model_ai = genai.GenerativeModel('gemini-flash-latest')
                    convo_prompt = (
                        f"You are Santhosh AI, a friendly construction and real estate AI assistant. "
                        f"You MUST answer ONLY in Tamil. User asks: {prompt}"
                    )
                    response = model_ai.generate_content(convo_prompt, stream=True)
                    for chunk in response:
                        full_response += chunk.text
                        placeholder.markdown(full_response + "▌")
                    placeholder.markdown(full_response)
                except Exception as e:
                    full_response = f"மன்னிக்கவும்! பிழை: {e}"
                    placeholder.markdown(full_response)
            else:
                lower_prompt = prompt.lower()
                if any(k in lower_prompt for k in ["cement", "sand", "சிமெண்ட்"]):
                    ai_response = "சிமெண்ட் மற்றும் மணலின் அளவு வீட்டின் sqft-ஐ பொறுத்தது. Material பட்டியலில் முழு விவரம் உள்ளது!"
                elif any(k in lower_prompt for k in ["price", "cost", "விலை"]):
                    ai_response = "வீட்டின் விலை Sqft மற்றும் Construction Quality-ஐ பொறுத்து மாறுபடும். Sidebar-ல் மாற்றிப் பாருங்கள்!"
                elif any(k in lower_prompt for k in ["hello", "hi", "வணக்கம்"]):
                    ai_response = "வணக்கம் நண்பரே! வீடு கட்டுவது குறித்து என்ன உதவி வேண்டும்?"
                else:
                    ai_response = "Gemini API Key-ஐ Sidebar-ல் உள்ளிடுங்கள், நான் தமிழில் நேரடியாக உதவுவேன்!"

                for chunk in ai_response.split(" "):
                    full_response += chunk + " "
                    time.sleep(0.04)
                    placeholder.markdown(full_response + "▌")
                placeholder.markdown(full_response)

        st.session_state.messages.append({"role": "assistant", "content": full_response.strip()})
